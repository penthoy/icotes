"""
Excel document handler

Supports reading and writing Excel files:
- .xlsx (via openpyxl/pandas)
- .xls (via xlrd - legacy)
- .xlsb (via pyxlsb - binary)
"""

import io
import logging
from typing import Any, Dict, List, Optional, Union

from .base import (
    DocumentFormat,
    DocumentHandler,
    HandlerResult,
    register_handler,
)

logger = logging.getLogger(__name__)


class ExcelHandler(DocumentHandler):
    """Handler for Excel files (.xlsx, .xls, .xlsb)"""
    
    @property
    def supported_formats(self) -> List[DocumentFormat]:
        return [
            DocumentFormat.EXCEL_XLSX,
            DocumentFormat.EXCEL_XLS,
            DocumentFormat.EXCEL_XLSB,
        ]
    
    async def read_content(
        self,
        file_data: bytes,
        file_path: str,
        options: Optional[Dict[str, Any]] = None
    ) -> HandlerResult:
        """
        Extract content from Excel file.
        
        Options:
            sheet_name: Specific sheet to read (default: all sheets)
            max_rows: Maximum rows per sheet (default: 1000)
            include_formulas: Include formula text (default: False)
            as_table: Return structured table data (default: True)
        """
        options = options or {}
        sheet_name = options.get("sheet_name")
        max_rows = options.get("max_rows", 1000)
        as_table = options.get("as_table", True)
        
        try:
            # Try pandas first (best for data extraction)
            return await self._read_with_pandas(
                file_data, file_path, sheet_name, max_rows, as_table
            )
        except ImportError:
            logger.warning("pandas not available, falling back to openpyxl")
            return await self._read_with_openpyxl(
                file_data, file_path, sheet_name, max_rows
            )
        except Exception as e:
            return HandlerResult(
                success=False,
                error=f"Failed to read Excel file: {str(e)}"
            )
    
    async def _read_with_pandas(
        self,
        file_data: bytes,
        file_path: str,
        sheet_name: Optional[str],
        max_rows: int,
        as_table: bool
    ) -> HandlerResult:
        """Read Excel using pandas for better data handling"""
        import pandas as pd
        
        # Determine engine based on file extension
        ext = file_path.lower().split('.')[-1]
        engine = {
            'xlsx': 'openpyxl',
            'xls': 'xlrd',
            'xlsb': 'pyxlsb'
        }.get(ext, 'openpyxl')
        
        try:
            # Read all sheets or specific sheet
            xlsx_file = io.BytesIO(file_data)
            
            if sheet_name:
                df_dict = {sheet_name: pd.read_excel(
                    xlsx_file, sheet_name=sheet_name, engine=engine, nrows=max_rows
                )}
            else:
                df_dict = pd.read_excel(
                    xlsx_file, sheet_name=None, engine=engine, nrows=max_rows
                )
            
            # Extract text content and tables
            content_parts = []
            tables = []
            
            for name, df in df_dict.items():
                content_parts.append(f"=== Sheet: {name} ===")
                content_parts.append(df.to_string(index=False, max_rows=max_rows))
                content_parts.append("")
                
                if as_table:
                    # Convert to structured format for AI
                    table_data = {
                        "sheet": name,
                        "columns": list(df.columns),
                        "row_count": len(df),
                        "data": df.head(max_rows).to_dict(orient="records")
                    }
                    tables.append(table_data)
            
            # Get metadata
            xlsx_file.seek(0)
            metadata = self._extract_metadata_pandas(xlsx_file, engine)
            
            return HandlerResult(
                success=True,
                content="\n".join(content_parts),
                metadata=metadata,
                tables=tables,
                sheets=list(df_dict.keys())
            )
            
        except ImportError as e:
            # Re-raise to fall back to openpyxl
            raise
        except Exception as e:
            logger.error(f"pandas Excel read error: {e}")
            raise
    
    def _extract_metadata_pandas(
        self, xlsx_file: io.BytesIO, engine: str
    ) -> Dict[str, Any]:
        """Extract workbook metadata"""
        metadata = {"engine": engine}
        
        if engine == "openpyxl":
            try:
                from openpyxl import load_workbook
                xlsx_file.seek(0)
                wb = load_workbook(xlsx_file, read_only=True, data_only=True)
                props = wb.properties
                if props:
                    metadata.update({
                        "title": props.title or "",
                        "author": props.creator or "",
                        "created": str(props.created) if props.created else "",
                        "modified": str(props.modified) if props.modified else "",
                    })
                wb.close()
            except Exception as e:
                logger.debug(f"Could not extract metadata: {e}")
        
        return metadata
    
    async def _read_with_openpyxl(
        self,
        file_data: bytes,
        file_path: str,
        sheet_name: Optional[str],
        max_rows: int
    ) -> HandlerResult:
        """Fallback: read Excel using openpyxl directly"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return HandlerResult(
                success=False,
                error="openpyxl not installed. Install with: pip install openpyxl"
            )
        
        xlsx_file = io.BytesIO(file_data)
        wb = load_workbook(xlsx_file, read_only=True, data_only=True)
        
        sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames
        content_parts = []
        tables = []
        
        for name in sheets_to_read:
            if name not in wb.sheetnames:
                continue
                
            ws = wb[name]
            content_parts.append(f"=== Sheet: {name} ===")
            
            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx >= max_rows:
                    content_parts.append(f"... (truncated at {max_rows} rows)")
                    break
                row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                content_parts.append(row_str)
                rows_data.append(list(row))
            
            content_parts.append("")
            
            # Create table structure
            if rows_data:
                headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows_data[0])]
                tables.append({
                    "sheet": name,
                    "columns": headers,
                    "row_count": len(rows_data) - 1,
                    "data": [
                        dict(zip(headers, [str(c) if c else "" for c in row]))
                        for row in rows_data[1:max_rows]
                    ]
                })
        
        # Extract metadata
        metadata = {}
        props = wb.properties
        if props:
            metadata = {
                "title": props.title or "",
                "author": props.creator or "",
                "created": str(props.created) if props.created else "",
                "modified": str(props.modified) if props.modified else "",
            }
        
        wb.close()
        
        return HandlerResult(
            success=True,
            content="\n".join(content_parts),
            metadata=metadata,
            tables=tables,
            sheets=sheets_to_read
        )
    
    async def write_content(
        self,
        content: Union[str, Dict[str, Any]],
        file_path: str,
        options: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Create Excel file from content.
        
        Content format:
            {
                "sheets": {
                    "Sheet1": {
                        "columns": ["A", "B", "C"],
                        "data": [{"A": 1, "B": 2, "C": 3}, ...]
                    }
                }
            }
        
        Options:
            sheet_name: Default sheet name if content is simple list
        """
        options = options or {}
        
        try:
            import pandas as pd
            from openpyxl import Workbook
        except ImportError:
            raise ImportError(
                "pandas and openpyxl required for Excel creation. "
                "Install with: pip install pandas openpyxl"
            )
        
        output = io.BytesIO()
        
        # Handle different content types
        if isinstance(content, str):
            # Parse CSV-like string into DataFrame
            df = pd.read_csv(io.StringIO(content))
            df.to_excel(output, index=False, sheet_name=options.get("sheet_name", "Sheet1"))
        elif isinstance(content, dict):
            if "sheets" in content:
                # Multiple sheets
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for sheet_name, sheet_data in content["sheets"].items():
                        df = pd.DataFrame(sheet_data.get("data", []))
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif "data" in content:
                # Single sheet with data
                df = pd.DataFrame(content["data"])
                df.to_excel(output, index=False, sheet_name=options.get("sheet_name", "Sheet1"))
            else:
                # Assume content is the data directly
                df = pd.DataFrame(content)
                df.to_excel(output, index=False, sheet_name=options.get("sheet_name", "Sheet1"))
        elif isinstance(content, list):
            # List of dicts
            df = pd.DataFrame(content)
            df.to_excel(output, index=False, sheet_name=options.get("sheet_name", "Sheet1"))
        else:
            raise ValueError(f"Unsupported content type: {type(content)}")
        
        output.seek(0)
        return output.read()


# Register handler when module is imported
_handler = ExcelHandler()
register_handler(_handler)
